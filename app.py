#region 1) 프로젝트 모듈 section.
"""
디스코드 모듈.
"""
import discord
from discord.ext import commands
from discord_buttons_plugin import *
"""

os 내장 모듈.
"""
import os
import datetime as dt
"""
엑셀  모듈.
"""
import xlsxwriter
import openpyxl
import boto3
#endregion
from pytz import timezone



"""
디스코드 커맨드 프리픽스, 클라이언트 버튼 쿼리 요청 초기화.
"""
client = commands.Bot(command_prefix ="!", intents=discord.Intents.all())
buttons = ButtonsClient(client)
#endregion

""" 
멀린 플레이어.
"""
import Merlin_player

cogs = [Merlin_player]
for i in range(len(cogs)):
      cogs[i].setup(client)

""" 
멀린 캘린더.
"""
import Merlin_calendar

cogs = [Merlin_calendar]
for i in range(len(cogs)):
      cogs[i].setup(client)
#endregion

#region 4) 멀린 봇 기동 이벤트 section. - 메인 백 윤정기.
"""
로컬 클라이언트 디스코드 봇 사용 준비 함수.
"""
@client.event
async def on_ready():
    # 봇을 온라인상태로 바꿔준다
    await client.change_presence(status=discord.Status.online, activity=None)
#endregion
aws_id = os.getenv("AWS_ACCESS_ID")
aws_key = os.getenv("AWS_ACCESS_KEY")
# Set AWS credentials 
s3 = boto3.client('s3', aws_access_key_id=f'{aws_id}',
    aws_secret_access_key=f'{aws_key}')

s3r = boto3.resource('s3', aws_access_key_id=f'{aws_id}',
    aws_secret_access_key=f'{aws_key}')
#endregion

#region 5) 버튼 이벤트 Section. - 1조 팀장 조시욱.
"""
버튼 클릭 이벤트 == 엑셀 파일 시트 이름 부여
"""
sheet_name_list = ['출근_기록부','직원_기록부','출입_대장부']
file_date = dt.datetime.now(timezone('Asia/Seoul')).strftime(f'%Y-%m-%d') # 파일 날짜.
filename = dt.datetime.now(timezone('Asia/Seoul')).strftime(f'%Y-%m-%d-{sheet_name_list[0]}') # 파일 이름.

# 출근 이벤트
@buttons.click
async def counter_start_button(ctx):
    company_name = f'{ctx.guild}'
    wt_calulator = dt.datetime.now(timezone('Asia/Seoul')).strftime('%H:%M')
    file_path = f'./commute_record_paper_folder/{company_name}-{filename}.xlsx'
    """ TODO : 
    1) company_name / 서버 이름 회사 이름 출력.
    2) wt_calulator / 출근 시간 계산.
    4) wt_employee_name / 출근 버튼 클릭한 직원이름 출력.
    """
    
    # 출근_기록부가 있을 때
    if os.path.exists(file_path):
        # 엑셀 파일을 읽어온다
        excel = openpyxl.load_workbook(file_path)
        excel_ws = excel['출근_기록부']
        emplo_ws = excel['직원_기록부']
        # 채널안의 맴버를 가져온다
        members = ctx.message.channel.members
        # 엑셀 파일 안의 맴버 정보를 모은다
        check_list = []
        for row in excel_ws.iter_rows(min_row=2, values_only=True):
            check_list.append(row[0])
        # 해당 엑셀 리스트에 유저가 있는지 확인
        if ctx.member.discriminator in check_list:
            pass
        else:
            # 리스트에 유저가 없다면 신입 사원 추가.
            for member in members:
                if ctx.member.discriminator == member.discriminator:
                    excel_ws.append([member.discriminator, member.display_name, '출근', f'{wt_calulator}'])
                    emplo_ws.append([member.discriminator, member.display_name])
                    excel.save(file_path)
                    await ctx.reply(f'"{member.display_name}"님이 출근 하였습니다.')
            # S3에 업로드
            s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
    
        for row in excel_ws.iter_rows(min_row=2):
              if row[0].value == ctx.member.discriminator and row[2].value == None:
                  # 0번쨰 셀들중에서 자신의 값과 같은 값이 있다면 해당셀의 이벤트 발생.
                  # if row[0].value == ctx.member.discriminator:
                      # 해당 row 에 접근하여 입력값을 넣어준다.
                row[2].value = '출근' # 근무 상태
                row[3].value = wt_calulator  # 현재 시간을 불러온다.
                # 출근할 때 마다 직원기록부 근무횟수 +1
                for row in emplo_ws.iter_rows(min_row=2):
                    if row[0].value == ctx.member.discriminator and row[2].value is not None:
                        row[2].value += 1
                    elif row[0].value == ctx.member.discriminator and row[2].value is None:
                        row[2].value = 1
                excel.save(file_path)
                # S3에 업로드
                s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
                # 출근 메세지 알림
                for member in members:
                    if ctx.member.discriminator == member.discriminator:
                        await ctx.reply(f'"{member.display_name}"님이 출근 하였습니다.')
                        break
              elif row[0].value == ctx.member.discriminator and row[2].value is not None:
                  for member in members:
                    if ctx.member.discriminator == member.discriminator:
                      await ctx.reply(f'"{member.display_name}"님은 이미 {row[2].value}처리 되었습니다.')
                      break
    # # 출근_기록부가 없을 때, 새로생성 (출근기록부, 직원기록부, 지급대장부, 출입대장부)
    else:
        await ctx.reply('출근 버튼을 새로 불러와 주세요.')
            
"""
퇴근 이벤트.
"""
@buttons.click
async def counter_end_button(ctx):
    company_name = f'{ctx.guild}'
    wt_calulator = dt.datetime.now(timezone('Asia/Seoul')).strftime('%H:%M')
    file_path = f'./commute_record_paper_folder/{company_name}-{filename}.xlsx'
    """ TODO : 
    1) company_name / 서버 이름 회사 이름 출력.
    2) wt_calulator / 출근 시간 계산.
    4) wt_employee_name / 출근 버튼 클릭한 직원이름 출력.
    """
    
    # 출근_기록부가 있을 때
    if os.path.exists(file_path):
        # 엑셀 파일을 읽어온다
        excel = openpyxl.load_workbook(file_path)
        excel_ws = excel['출근_기록부']
        emplo_ws = excel['직원_기록부']
        # 채널안의 맴버를 가져온다
        members = ctx.message.channel.members
        # 엑셀 파일 안의 맴버 정보를 모은다
        check_list = []
        for row in excel_ws.iter_rows(min_row=2, values_only=True):
            check_list.append(row[0])
        # 해당 엑셀 리스트에 유저가 있는지 확인
        if ctx.member.discriminator in check_list:
            pass
        else:
            # 리스트에 유저가 없다면 신입 사원 추가멘트.
            await ctx.reply(f'"{member.display_name}"님은 사원정보가 없습니다.')
            
        for row in excel_ws.iter_rows(min_row=2):
              if row[0].value == ctx.member.discriminator and row[2].value == '출근'and row[4].value == None:
                  # 0번쨰 셀들중에서 자신의 값과 같은 값이 있다면 해당셀의 이벤트 발생.
                  # if row[0].value == ctx.member.discriminator:
                      # 해당 row 에 접근하여 입력값을 넣어준다.
                row[4].value = wt_calulator  # 현재 시간을 불러온다.
                # 퇴근시간과 출근시간의 차 를 구한다
                intvl = (int(row[4].value[:2]) * 60 + int(row[4].value[3:5])) - (int(row[3].value[:2]) * 60 + int(row[3].value[3:5]))
                if intvl >= 0:
                    hint = intvl // 60
                else:
                    hint = (intvl // 60) + 24
                mint = intvl % 60
                # 근무한 시간 표기
                row[5].value = f'{hint}:{mint}'
                # 직원기록부에도 누적 근무시간 표기
                for row in emplo_ws.iter_rows(min_row=2):
                    if row[0].value == ctx.member.discriminator:
                        if row[3].value is not None:
                            achint = int(row[3].value.split(':')[0]) + hint
                            acmint =  int(row[3].value.split(':')[1]) + mint
                            row[3].value = f'{achint}:{acmint}'
                        else:
                            row[3].value = f'{hint}:{mint}'
                excel.save(file_path)
                # S3에 업로드
                s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
                # 퇴근 메세지 전송
                for member in members:
                    if ctx.member.discriminator == member.discriminator:
                        await ctx.reply(f'"{member.display_name}"님이 퇴근 하였습니다.')
                        break
              elif row[0].value == ctx.member.discriminator and row[4].value is not None and row[2].value == '출근':
                  for member in members:
                    if ctx.member.discriminator == member.discriminator:
                      await ctx.reply(f'"{member.display_name}"님은 이미 퇴근 하였습니다.')
                      break
                  
    # 출근_기록부가 없을 때, 새로생성 (출근기록부, 직원기록부, 지급대장부, 출입대장부)
    else:
        await ctx.reply('출근 버튼을 새로 불러와 주세요.')
"""병결 이벤트.
"""
@buttons.click
async def counter_illness_button(ctx):
    company_name = f'{ctx.guild}'
    wt_calulator = dt.datetime.now(timezone('Asia/Seoul')).strftime('%H:%M')
    file_path = f'./commute_record_paper_folder/{company_name}-{filename}.xlsx'
    """ TODO : 
    1) company_name / 서버 이름 회사 이름 출력.
    2) wt_calulator / 출근 시간 계산.
    4) wt_employee_name / 출근 버튼 클릭한 직원이름 출력.
    """
    
    # 출근_기록부가 있을 때
    if os.path.exists(file_path):
        # 엑셀 파일을 읽어온다
        excel = openpyxl.load_workbook(file_path)
        excel_ws = excel['출근_기록부']
        # 채널안의 맴버를 가져온다
        members = ctx.message.channel.members
        # 엑셀 파일 안의 맴버 정보를 모은다
        check_list = []
        for row in excel_ws.iter_rows(min_row=2, values_only=True):
            check_list.append(row[0])
        # 해당 엑셀 리스트에 유저가 있는지 확인
        if ctx.member.discriminator in check_list:
            pass
        else:
            # 리스트에 유저가 없다면 신입 사원 추가.
            for member in members:
                if ctx.member.discriminator == member.discriminator:
                    excel_ws.append([member.discriminator, member.display_name, '병결', f'{wt_calulator}', '-', '-'])
                    excel.save(file_path)
                    await ctx.reply(f'"{member.display_name}"님은 병결 입니다.')
            # S3에 업로드
            s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
    
        for row in excel_ws.iter_rows(min_row=2):
              if row[0].value == ctx.member.discriminator and row[2].value == None:
                  # 0번쨰 셀들중에서 자신의 값과 같은 값이 있다면 해당셀의 이벤트 발생.
                  # if row[0].value == ctx.member.discriminator:
                      # 해당 row 에 접근하여 입력값을 넣어준다.
                row[2].value = '병결' # 근무 상태
                row[3].value = '-'
                excel.save(file_path)
                # S3에 업로드
                s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
                # 병결 알림 메세지
                for member in members:
                    if ctx.member.discriminator == member.discriminator:
                        await ctx.reply(f'"{member.display_name}"님은 병결 입니다.')
                        break
              elif row[0].value == ctx.member.discriminator and row[2].value is not None:
                  for member in members:
                    if ctx.member.discriminator == member.discriminator:
                      await ctx.reply(f'"{member.display_name}"님은 이미 {row[2].value}처리 되었습니다.')
                      break
    # # 출근_기록부가 없을 때, 새로생성 (출근기록부, 직원기록부, 지급대장부, 출입대장부)
    else:
        await ctx.reply('출근 버튼을 새로 불러와 주세요.')
    
"""외근 이벤트.
"""
@buttons.click
async def out_on_business(ctx):
    company_name = f'{ctx.guild}'
    wt_calulator = dt.datetime.now(timezone('Asia/Seoul')).strftime('%H:%M')
    file_path = f'./commute_record_paper_folder/{company_name}-{filename}.xlsx'
    
    """ TODO : 
    1) company_name / 서버 이름 회사 이름 출력.
    2) wt_calulator / 출근 시간 계산.
    4) wt_employee_name / 출근 버튼 클릭한 직원이름 출력.
    """
    
    # 출근_기록부가 있을 때
    if os.path.exists(file_path):
        # 엑셀 파일을 읽어온다
        excel = openpyxl.load_workbook(file_path)
        excel_ws = excel['출근_기록부']
        emplo_ws = excel['직원_기록부']
        # 채널안의 맴버를 가져온다
        members = ctx.message.channel.members
        # 엑셀 파일 안의 맴버 정보를 모은다
        check_list = []
        for row in excel_ws.iter_rows(min_row=2, values_only=True):
            check_list.append(row[0])
        # 해당 엑셀 리스트에 유저가 있는지 확인
        if ctx.member.discriminator in check_list:
            pass
        else:
            # 리스트에 유저가 없다면 신입 사원 추가.
            for member in members:
                if ctx.member.discriminator == member.discriminator:
                    excel_ws.append([member.discriminator, member.display_name, '외근', f'{wt_calulator}', '-', '-'])
                    excel.save(file_path)
                    await ctx.reply(f'"{member.display_name}"님은 외근 입니다.')
            # S3에 업로드
            s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
    
        for row in excel_ws.iter_rows(min_row=2):
              if row[0].value == ctx.member.discriminator and row[2].value == None:
                  # 0번쨰 셀들중에서 자신의 값과 같은 값이 있다면 해당셀의 이벤트 발생.
                  # if row[0].value == ctx.member.discriminator:
                      # 해당 row 에 접근하여 입력값을 넣어준다.
                row[2].value = '외근' # 근무 상태
                row[3].value = wt_calulator  # 현재 시간을 불러온다.
                for row in emplo_ws.iter_rows(min_row=2):
                    if row[0].value == ctx.member.discriminator and row[2].value is not None:
                        row[2].value += 1
                    elif row[0].value == ctx.member.discriminator and row[2].value is None:
                        row[2].value = 1
                excel.save(file_path)
                # S3에 업로드
                s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
                for member in members:
                    if ctx.member.discriminator == member.discriminator:
                        await ctx.reply(f'"{member.display_name}"님은 외근 입니다.')
                        break
              elif row[0].value == ctx.member.discriminator and row[2].value is not None:
                  for member in members:
                    if ctx.member.discriminator == member.discriminator:
                      await ctx.reply(f'"{member.display_name}"님은 이미 {row[2].value}처리 되었습니다.')
                      break
    # # 출근_기록부가 없을 때, 새로생성 (출근기록부, 직원기록부, 지급대장부, 출입대장부)
    else:
        await ctx.reply('출근 버튼을 새로 불러와 주세요.')
#endregion

#region 6) 멀린 봇 서버에 초대 링크 임베디드 생성 커맨드 section. - 1조 팀장 조시욱.
"""
초대 링크 임베디드 생성 커맨드.
"""
@client.command()
async def counter(ctx):
    #region 1) 엠베드  initializer section.
    
    embed = discord.Embed(title=f"Time Recorder", color = 0x2ba191,
                          description=f"{file_date} 일자 출퇴근 기록부ㅤㅤㅤㅤㅤㅤㅤㅤㅤㅤ")
    
    await buttons.send(
        content = None,
        embed = embed,
        channel = ctx.channel.id,
        components = [
            ActionRow([
                # 출근 버튼.
                Button(
                    label = "출근",
                    style = ButtonType().Primary,
                    custom_id = "counter_start_button",
                    emoji = {
                            "id": None,
                            "name": "☕",
                            "animated": False
                    },
                ),
                # 퇴근 버튼.
                Button(
                    style = ButtonType().Success,
                    label = "퇴근",
                    custom_id = "counter_end_button",
                    emoji = {
                            "id": None,
                            "name": "🌙",
                            "animated": False
                    },
                ),
                # 병결 버튼.
                Button(
                    style = ButtonType().Danger,
                    label = "병결",
                    custom_id = "counter_illness_button",
                    emoji = {
                            "id": None,
                            "name": "🤒",
                            "animated": False
                    },
                ),
                # 외근 버튼.
                Button(
                    style = ButtonType().Secondary,
                    label = "외근",
                    custom_id = "out_on_business",
                    emoji = {
                            "id": None,
                            "name": "💼",
                            "animated": False
                    },
                ),
            ])
        ]
    )
    """
    1) company_name / 서버 이름 회사 이름 출력.
    """
    company_name = f'{ctx.guild}'
    file_path = f'./commute_record_paper_folder/{company_name}-{filename}.xlsx'

    # 오늘자 출근_기록부가 있을 때
    if os.path.exists(file_path):
        pass
    # 오늘자 출근_기록부가 없을 때, 새로생성 (출근기록부, 직원기록부, 지급대장부, 출입대장부)
    else:
        # 이전 엑셀 파일이 있는지 확인
        files_Path = "./commute_record_paper_folder/" # 파일들이 들어있는 폴더
        file_name_and_time_lst = []
        # 해당 경로에 있는 파일들의 생성시간을 함께 리스트로 넣어줌. 
        for f_name in os.listdir(f"{files_Path}"):
            written_time = os.path.getctime(f"{files_Path}{f_name}")
            file_name_and_time_lst.append((f_name, written_time))
            
        # 이전 엑셀파일이 하나라도 있다면
        if len(file_name_and_time_lst) > 1:
            # 생성시간 역순으로 정렬하고, 
            sorted_file_lst = sorted(file_name_and_time_lst, key=lambda x: x[1], reverse=True)
            # 가장 앞에 있는 파일이 최신파일
            recent_file_name = sorted_file_lst[0][0]
            recent_file_path = f'./commute_record_paper_folder/{recent_file_name}'
            
            # 파일이름, 워크북 옵션(대용량 파일을 효율적으로 쓸 수 있도록 메모리에 저장되는 데이터의 양을 줄입니다.)
            with xlsxwriter.Workbook(f'./commute_record_paper_folder/{company_name}-{filename}.xlsx', {'constant_memory': True}) as workbook:
                # 0) 엑셀 파일을 실행하면 읽기 전용으로 할지 안할지 알러트가 뜨게 한다.
                workbook.read_only_recommended()
                counter_record_sheet = workbook.add_worksheet(sheet_name_list[0]) # '출근_기록부'
                employee_register_sheet = workbook.add_worksheet(sheet_name_list[1]) # '직원_기록부'
                entry_ticket_sheet = workbook.add_worksheet(sheet_name_list[2]) # '출입 대장부'
                
                # 1) 카테고리 텍스트에 bold 포맷이 적용된다.
                # category_text_bolding = workbook.add_format({'bold': True, 'italic': True}) # 이탈릭체가 적용된다.
                category_text_bolding = workbook.add_format({'bold': True})
                counter_record_sheet.write('A1', '식별 번호', category_text_bolding)
                counter_record_sheet.write('B1', '직원 이름', category_text_bolding)
                counter_record_sheet.write('C1', '출근 상태', category_text_bolding)
                counter_record_sheet.write('D1', '출근 시간', category_text_bolding)
                counter_record_sheet.write('E1', '퇴근 시간', category_text_bolding)
                counter_record_sheet.write('F1', '근무 시간', category_text_bolding)

                entry_ticket_sheet.write('A1', '방문 일자', category_text_bolding)
                entry_ticket_sheet.write('B1', '방문 시간', category_text_bolding)
                entry_ticket_sheet.write('C1', '방문자', category_text_bolding)
                entry_ticket_sheet.write('D1', '출입목적', category_text_bolding)
                entry_ticket_sheet.write('E1', '담당자', category_text_bolding) 
                # 2) 서버 멤버를 찾아서 memebers 에 저장한다.
                members = ctx.message.channel.members
                # 3) 출근 히스토리 
                clock_List = list()
                # 4) 서버내의 봇의 
                for member in members:
                # 만약 멤버가 봇이 아니라면 출근 리스트에 추가되도록 한다.
                    if member.bot == False :
                        # 멤버들은 해당 셀 칸에 멤버 식별자, 멤버 닉네임 또는 이름 추가.
                        clock_List.append([member.discriminator, member.display_name, '', '', '', ''])
                row = 1
                col = 0
                # 5) 출근자, 츨근상태, 출근 시간을 한번에 기록한다.
                for employee_rank, employee_name, employee_work_status, employee_counter_start_time, employee_counter_end_time, employee_daily_wage in (clock_List):
                    counter_record_sheet.write(row, col, employee_rank)
                    counter_record_sheet.write(row, col + 1, employee_name)
                    counter_record_sheet.write(row, col + 2, employee_work_status)
                    counter_record_sheet.write(row, col + 3, employee_counter_start_time)
                    counter_record_sheet.write(row, col + 4, employee_counter_end_time)
                    counter_record_sheet.write(row, col + 5, employee_daily_wage)
                    row += 1
                    
            # 화일간 시트 복사하기
            wb1 = openpyxl.load_workbook(recent_file_path)
            ws1 = wb1["직원_기록부"]
            wb2 = openpyxl.load_workbook(file_path)
            ws2 = wb2["직원_기록부"]
            for row in ws1:
                for cell in row:
                    ws2[cell.coordinate].value = cell.value
            wb2.save(file_path)
            # S3에 업로드
            s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
        
        # 엑셀파일이 없으면 새롭게 생성
        else:
            # 파일이름, 워크북 옵션(대용량 파일을 효율적으로 쓸 수 있도록 메모리에 저장되는 데이터의 양을 줄입니다.)
            with xlsxwriter.Workbook(f'./commute_record_paper_folder/{company_name}-{filename}.xlsx', {'constant_memory': True}) as workbook:
                # 0) 엑셀 파일을 실행하면 읽기 전용으로 할지 안할지 알러트가 뜨게 한다.
                workbook.read_only_recommended()
                counter_record_sheet = workbook.add_worksheet(sheet_name_list[0]) # '출근_기록부'
                employee_register_sheet = workbook.add_worksheet(sheet_name_list[1]) # '직원_기록부'
                entry_ticket_sheet = workbook.add_worksheet(sheet_name_list[2]) # '출입 대장부'
                # 1) 카테고리 텍스트에 bold 포맷이 적용된다.
                # category_text_bolding = workbook.add_format({'bold': True, 'italic': True}) # 이탈릭체가 적용된다.
                category_text_bolding = workbook.add_format({'bold': True})
                counter_record_sheet.write('A1', '식별 번호', category_text_bolding)
                counter_record_sheet.write('B1', '직원 이름', category_text_bolding)
                counter_record_sheet.write('C1', '출근 상태', category_text_bolding)
                counter_record_sheet.write('D1', '출근 시간', category_text_bolding)
                counter_record_sheet.write('E1', '퇴근 시간', category_text_bolding)
                counter_record_sheet.write('F1', '근무 시간', category_text_bolding)
                
                entry_ticket_sheet.write('A1', '방문 일자', category_text_bolding)
                entry_ticket_sheet.write('B1', '방문 시간', category_text_bolding)
                entry_ticket_sheet.write('C1', '방문자', category_text_bolding)
                entry_ticket_sheet.write('D1', '출입목적', category_text_bolding)
                entry_ticket_sheet.write('E1', '담당자', category_text_bolding)
                                
                employee_register_sheet.write('A1', '식별 번호', category_text_bolding)
                employee_register_sheet.write('B1', '직원 이름', category_text_bolding)
                employee_register_sheet.write('C1', '근무 횟수', category_text_bolding)
                employee_register_sheet.write('D1', '총 근무 시간', category_text_bolding)
                employee_register_sheet.write('E1', '여분 월차', category_text_bolding)
                # 2) 서버 멤버를 찾아서 memebers 에 저장한다.
                members = ctx.message.channel.members
                # 3) 출근 히스토리 
                clock_List = list()
                # 4) 서버내의 봇의 
                for member in members:
                # 만약 멤버가 봇이 아니라면 출근 리스트에 추가되도록 한다.
                    if member.bot == False :
                        # 멤버들은 해당 셀 칸에 멤버 식별자, 멤버 닉네임 또는 이름 추가.
                        clock_List.append([member.discriminator, member.display_name, '', '', '', ''])
                row = 1
                col = 0
                # 5) 고유번호, 출근자를 기록한다.
                for employee_rank, employee_name, employee_work_status, employee_counter_start_time, employee_counter_end_time, employee_daily_wage in (clock_List):
                    counter_record_sheet.write(row, col, employee_rank)
                    counter_record_sheet.write(row, col + 1, employee_name)
                    counter_record_sheet.write(row, col + 2, employee_work_status)
                    counter_record_sheet.write(row, col + 3, employee_counter_start_time)
                    counter_record_sheet.write(row, col + 4, employee_counter_end_time)
                    counter_record_sheet.write(row, col + 5, employee_daily_wage)
                    
                    employee_register_sheet.write(row, col, employee_rank)
                    employee_register_sheet.write(row, col + 1, employee_name)
                    employee_register_sheet.write(row, col + 2, employee_work_status)
                    employee_register_sheet.write(row, col + 3, employee_counter_start_time)
                    employee_register_sheet.write(row, col + 4, employee_counter_end_time)
                    employee_register_sheet.write(row, col + 5, employee_daily_wage)
                    row += 1
                    
            s3.upload_file(f'{file_path}', 'merlin-bucket', f'commute_record_paper_folder/{company_name}-{filename}.xlsx')
#endregion

#region 6) 멀린 봇 서버에 초대 링크 임베디드 생성 커맨드 section. - 1조 팀장 조시욱.
"""
초대 링크 임베디드 생성 커맨드.
"""
@client.command()
async def invite(ctx):  
    #region 1) 엠베드  initializer section.
    
    embed = discord.Embed(title=f" Invite Merlin Bot to your server !", color = 0x2ba191,
                          description=f"Merlin Bot is easy to use and included powerful commands ERP Program.\
                                        some commands gonna make you access excel files and downloads on your any devices.\
                                        officially Merlin Bot is not stablized to use so we still keep updating at our server.\
                                        Do you Wanna try this Merlin Bot test program ? then invite Link here !\
                                        [click](https://discord.com/api/oauth2/authorize?client_id={client.user.id}&permissions=8&scope=applications.commands%20bot).")
    # 엠베딩 푸터 라인.
    embed.set_footer(text="Information requested by: {0}".format(ctx.author.display_name))
    #endregion
    
    await buttons.send(
        content = None,
        embed = embed,
        channel = ctx.channel.id,
        components = [
            ActionRow([
                # 클라이언트 링크 서버에 등록 버튼.
                Button(
                    style = ButtonType().Link,
                    label = "Invite",
                    url ="https://discord.com/api/oauth2/authorize?client_id=950766027535421460&permissions=8&scope=applications.commands%20bot"
                ),
                
            ])
        ]
    )


# 출근파일 다운로드
@client.command()
async def download(ctx):
    company_name = f'{ctx.guild}'
    bucket = s3r.Bucket('merlin-bucket')
    file_list = ['https://merlin-bucket.s3.amazonaws.com/' + obj.key for obj in bucket.objects.all() if obj.key.find(f'{company_name}')>-1]
    
    embed = discord.Embed(title="Time Recorder", color = 0x2ba191,
                          description=f"{file_list}")
    
    await ctx.send(embed=embed)
    
      
"""
로컬 클라이언트 디스코드 실행부.
"""
# 로컬 클라이언트 디스코드 프로그램에 이 봇을 연결해준다.
token = os.getenv("TOKEN")
client.run(token)
